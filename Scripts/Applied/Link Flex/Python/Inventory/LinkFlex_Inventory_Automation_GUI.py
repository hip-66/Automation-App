"""
LinkFlex Inventory Automation - GUI Edition

Reads the ATP_LinkFlex_Firewall*.txt report and queries the Proxmox and
iDRAC (Redfish) APIs to populate "LinkFlex Inventory.xlsx" automatically.
"""

import os
import sys
import re
import json
import glob
import shutil
import queue
import threading
import traceback
import subprocess
from datetime import datetime

# ============================================================
# Dependency bootstrap (runs before the heavier third-party imports)
# ============================================================
def _install_and_import(package, import_name=None):
    import_name = import_name or package
    try:
        __import__(import_name)
    except ImportError:
        print(f"[INIT] Installing missing dependency: {package} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", package])

for _pkg, _imp in (("openpyxl", "openpyxl"), ("requests", "requests"), ("urllib3", "urllib3")):
    _install_and_import(_pkg, _imp)

import openpyxl
import requests
import urllib3

# Tkinter is only needed for the manual/standalone GUI path - imported
# lazily-but-safely here (not required, not fatal if missing) so headless
# mode (see run_headless() / main()) still works on a machine without
# tcl/tk, which is exactly the kind of constrained VM this app runs on.
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    _TKINTER_AVAILABLE = True
except ImportError:
    _TKINTER_AVAILABLE = False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================
# Excel layout constants (sheet "RAW Data" in LinkFlex Inventory.xlsx)
# ============================================================
NAME_COL = 2         # Column B - row label / disk slot number
VAL_COL_FW = 3        # Column C - Firewall MAC / Disk size (GB)
VAL_COL_DISK_SN = 4   # Column D - Disk serial number
VAL_COL_VM_1 = 6      # Column F - VM MAC (Data)
VAL_COL_VM_2 = 7      # Column G - VM MAC (MGMT)

FIREWALL_GLOB = "ATP_LinkFlex_Firewall*.txt"
DEFAULT_EXCEL_NAME = "LinkFlex Inventory.xlsx"
COMPLETED_EXPORT_SUBDIR = "Completed Inventory"

# Firewall report key -> Excel row label under the "FIREWALL" section
TARGET_FW_PORTS = {
    'port1': 'FW MAC PORT 1', 'port2': 'FW MAC PORT 2', 'port3': 'FW MAC PORT 3',
    'port4': 'FW MAC PORT 4', 'port5': 'FW MAC PORT 5', 'port6': 'FW MAC PORT 6',
    'port7': 'FW MAC PORT 7', 'port8': 'FW MAC PORT 8', 'port9': 'FW MAC PORT 9',
    'port10': 'FW MAC PORT 10', 'port11': 'FW MAC PORT 11', 'port12': 'FW MAC PORT 12',
    'port15': 'FW MAC PORT 15', 'port16': 'FW MAC PORT 16',
    'x1': 'FW MAC PORT x1', 'x2': 'FW MAC PORT x2', 'x3': 'FW MAC PORT x3',
    'x6': 'FW MAC PORT x6', 'x8': 'FW MAC PORT x8',
}

# Parsed iDRAC NIC key -> header keywords used to locate the target column
IDRAC_NIC_COLUMN_KEYWORDS = {
    "emb_1": ["NIC1", "Emb"], "emb_2": ["NIC2", "Emb"],
    "int_1": ["NIC1", "Int"], "int_2": ["NIC2", "Int"],
    "int_3": ["NIC3", "Int"], "int_4": ["NIC4", "Int"],
}

DISK_SIZE_TOLERANCE_GB = 10.0  # accounts for GiB/GB rounding differences between iDRAC and the sheet


# ============================================================
# Small helpers
# ============================================================
def normalize_text(text):
    if not text:
        return ""
    return str(text).lower().replace(" ", "").strip()


def normalize_url(url):
    url = (url or "").strip()
    if not url.startswith("http"):
        url = "https://" + url
    return url.rstrip("/")


def write_safe(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    if type(cell).__name__ == "MergedCell":
        return False
    cell.value = value
    return True


def find_column_index(sheet, row_idx, keywords):
    for col in range(1, sheet.max_column + 1):
        val = normalize_text(sheet.cell(row=row_idx, column=col).value)
        if all(k.lower() in val for k in keywords):
            return col
    return None


def find_firewall_file(folder, log):
    files = glob.glob(os.path.join(folder, FIREWALL_GLOB))
    if not files:
        log(f"[ERROR] No Firewall report found (pattern: {FIREWALL_GLOB}) in the selected folder.")
        return None
    latest = max(files, key=os.path.getmtime)
    log(f"[INFO] Firewall report: {os.path.basename(latest)}")
    return latest


def find_excel_file(folder, log):
    exact = os.path.join(folder, DEFAULT_EXCEL_NAME)
    if os.path.exists(exact):
        log(f"[INFO] Excel file: {DEFAULT_EXCEL_NAME}")
        return exact

    candidates = sorted(
        f for f in glob.glob(os.path.join(folder, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")  # skip Excel lock files
    )
    if not candidates:
        log("[FATAL] No .xlsx file found in the selected folder.")
        return None
    if len(candidates) > 1:
        log(f"[WARNING] Multiple .xlsx files found; using '{os.path.basename(candidates[0])}'.")
    else:
        log(f"[INFO] Excel file: {os.path.basename(candidates[0])}")
    return candidates[0]


def load_workbook_safe(path, log, confirm):
    while True:
        try:
            wb = openpyxl.load_workbook(path)
            return wb, wb.active
        except PermissionError:
            log(f"[ERROR] Cannot open '{os.path.basename(path)}' - permission denied.")
            if not confirm(
                "Excel File Locked",
                f"Could not open:\n{path}\n\n"
                "The file may be open in Excel. Close it, then click Retry.",
            ):
                return None, None
        except Exception as e:
            log(f"[FATAL] Failed to load workbook: {e}")
            return None, None


def save_excel_safe(workbook, path, log, confirm):
    while True:
        try:
            workbook.save(path)
            log(f"[SUCCESS] Saved '{os.path.basename(path)}'.")
            return True
        except PermissionError:
            log(f"[ERROR] '{os.path.basename(path)}' is open in another program.")
            if not confirm(
                "Excel File Locked",
                f"Could not save:\n{path}\n\n"
                "The file appears to be open (e.g. in Excel).\n"
                "Close it, then click Retry.",
            ):
                log("[WARNING] Save skipped by user - changes remain unsaved for this step.")
                return False
        except Exception as e:
            log(f"[ERROR] Save failed: {e}")
            return False


def export_completed_copy(excel_path, log):
    """Copy the finished workbook into a timestamped 'Completed Inventory' subfolder
    so the filled-in result is preserved even if the source file gets reused/overwritten later."""
    try:
        folder = os.path.dirname(excel_path)
        out_dir = os.path.join(folder, COMPLETED_EXPORT_SUBDIR)
        os.makedirs(out_dir, exist_ok=True)

        base = os.path.splitext(os.path.basename(excel_path))[0]
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        dest = os.path.join(out_dir, f"{base}_COMPLETED_{timestamp}.xlsx")

        shutil.copy2(excel_path, dest)
        log(f"[SUCCESS] Exported a full completed copy -> {dest}")
        return dest
    except Exception as e:
        log(f"[WARNING] Could not export a completed copy: {e}")
        return None


# ============================================================
# FIREWALL
# ============================================================
def process_firewall(folder_path, ws, log):
    log("\n--- Processing Firewall ---")
    fw_file = find_firewall_file(folder_path, log)
    if not fw_file:
        return

    mac_mapping = {}
    serial_number = "N/A"

    try:
        with open(fw_file, "r", encoding="utf-8") as f:
            content = f.read()
        if "----- INVENTORY -----" in content:
            inv_section = content.split("----- INVENTORY -----")[1]

            serial_match = re.search(r"Serial Number:\s*([A-Za-z0-9-]+)", inv_section)
            if serial_match:
                serial_number = serial_match.group(1).strip()

            for line in inv_section.splitlines():
                match = re.search(r"^\s*([a-zA-Z0-9]+):\s*([0-9a-fA-F:]+)", line)
                if match:
                    mac_mapping[match.group(1)] = match.group(2).lower()
        else:
            log("[WARNING] Firewall report has no '----- INVENTORY -----' section.")
    except Exception as e:
        log(f"[ERROR] Reading Firewall file failed: {e}")
        return

    norm_targets = {k: normalize_text(v) for k, v in TARGET_FW_PORTS.items()}

    for row in range(1, ws.max_row + 1):
        cell_obj = ws.cell(row=row, column=NAME_COL)
        if type(cell_obj).__name__ == "MergedCell":
            continue

        val = normalize_text(cell_obj.value)

        if val == "fw":
            write_safe(ws, row, VAL_COL_FW, serial_number)
            log(f"[UPDATE] Firewall Serial Number -> {serial_number}")
            continue

        for key, target in norm_targets.items():
            if target == val:
                mac = mac_mapping.get(key, "N/A")
                write_safe(ws, row, VAL_COL_FW, mac)
                log(f"[UPDATE] {TARGET_FW_PORTS[key]} -> {mac}")


# ============================================================
# PROXMOX API
# ============================================================
def process_proxmox_api(ws, cfg, log):
    log("\n--- Processing Proxmox (via API) ---")
    url = normalize_url(cfg["prox_url"])

    user = cfg["prox_user"]
    if "@" not in user:
        user += "@pam"

    log("[API] Authenticating to Proxmox...")
    try:
        resp = requests.post(
            f"{url}/api2/json/access/ticket",
            data={"username": user, "password": cfg["prox_pass"]},
            verify=False,
            timeout=10,
        )
        resp.raise_for_status()
        auth_data = resp.json()["data"]
        headers = {
            "CSRFPreventionToken": auth_data["CSRFPreventionToken"],
            "Cookie": f"PVEAuthCookie={auth_data['ticket']}",
        }
        log("[SUCCESS] Connected to Proxmox API.")
    except Exception:
        log("[ERROR] Proxmox API login failed (check credentials/URL).")
        return

    log("[API] Fetching VM list...")
    try:
        res_resp = requests.get(
            f"{url}/api2/json/cluster/resources?type=vm", headers=headers, verify=False, timeout=10
        )
        res_resp.raise_for_status()
        vms = res_resp.json()["data"]
        vm_dict = {str(vm.get("name", "")).lower(): vm for vm in vms}
    except Exception as e:
        log(f"[ERROR] Failed to fetch VMs: {e}")
        return

    start_row = -1
    for row in range(1, ws.max_row + 1):
        if "PVE VM's" in str(ws.cell(row=row, column=NAME_COL).value):
            start_row = row + 2
            break
    if start_row == -1:
        log("[ERROR] 'PVE VM's' section not found in the Excel sheet.")
        return

    end_row = ws.max_row
    for row in range(start_row, ws.max_row + 1):
        if "Hypervisor Servers" in str(ws.cell(row=row, column=NAME_COL).value):
            end_row = row - 1
            break

    for row in range(start_row, end_row + 1):
        name_cell = ws.cell(row=row, column=NAME_COL)
        if type(name_cell).__name__ == "MergedCell":
            continue

        vm_name_raw = name_cell.value
        if not vm_name_raw or str(vm_name_raw).lower() == "none":
            continue

        vm_name = str(vm_name_raw).lower()
        dest_cell = ws.cell(row=row, column=VAL_COL_VM_1)
        if type(dest_cell).__name__ == "MergedCell":
            continue

        if vm_name in vm_dict:
            node = vm_dict[vm_name]["node"]
            vmid = vm_dict[vm_name]["vmid"]
            log(f"[CHECK] Querying VM: {vm_name_raw} (ID: {vmid})")

            macs = []
            try:
                conf_resp = requests.get(
                    f"{url}/api2/json/nodes/{node}/qemu/{vmid}/config", headers=headers, verify=False, timeout=5
                )
                if conf_resp.status_code == 200:
                    vm_config = conf_resp.json().get("data", {})
                    for i in range(4):
                        net_str = vm_config.get(f"net{i}", "")
                        match = re.search(r"([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})", net_str)
                        if match:
                            macs.append(match.group(0).lower())
            except Exception as e:
                log(f"[WARNING] API error for VM {vm_name_raw}: {e}")

            if len(macs) == 0:
                write_safe(ws, row, VAL_COL_VM_1, "N/A")
                log(f"[UPDATE] {vm_name_raw} -> N/A")
            elif len(macs) == 1:
                write_safe(ws, row, VAL_COL_VM_1, macs[0])
                log(f"[UPDATE] {vm_name_raw} -> {macs[0]}")
            else:
                mac_data = f"Data - {macs[0]}"
                mac_mgmt = f"MGMT - {macs[1]}"
                write_safe(ws, row, VAL_COL_VM_1, mac_data)
                write_safe(ws, row, VAL_COL_VM_2, mac_mgmt)
                log(f"[UPDATE] {vm_name_raw} -> {mac_data} | {mac_mgmt}")
        else:
            if len(vm_name) > 2 and "slots" not in vm_name:
                write_safe(ws, row, VAL_COL_VM_1, "N/A")
                log(f"[UPDATE] {vm_name_raw} -> N/A (no matching VM on Proxmox)")


# ============================================================
# iDRAC API (Redfish)
# ============================================================
def process_idrac_api(ws, cfg, log):
    log("\n--- Processing iDRAC (via Redfish API) ---")
    url = normalize_url(cfg["idrac_url"])
    auth = (cfg["idrac_user"], cfg["idrac_pass"])
    server_name = cfg["server_name"]

    # ---------------------------------------------------------
    # PART 1: NETWORK INTERFACES (NICs)
    # ---------------------------------------------------------
    header_row = -1
    for row in range(1, ws.max_row + 1):
        if "Hypervisor Servers" in str(ws.cell(row=row, column=NAME_COL).value):
            header_row = row + 1
            break

    target_row = -1
    if header_row != -1:
        for row in range(header_row + 1, ws.max_row + 1):
            if normalize_text(ws.cell(row=row, column=NAME_COL).value) == normalize_text(server_name):
                target_row = row
                break

    if target_row == -1:
        log(f"[ERROR] Server '{server_name}' not found in the Hypervisor Servers section.")
    else:
        log("[API] Fetching network data from iDRAC...")
        mac_data = {}
        failed = 0
        try:
            resp = requests.get(
                f"{url}/redfish/v1/Systems/System.Embedded.1/EthernetInterfaces", auth=auth, verify=False, timeout=10
            )
            resp.raise_for_status()
            interfaces = resp.json().get("Members", [])

            emb_list, int_list = [], []
            for iface in interfaces:
                iface_url = iface.get("@odata.id")
                if not iface_url:
                    continue
                try:
                    iface_resp = requests.get(f"{url}{iface_url}", auth=auth, verify=False, timeout=5)
                    if iface_resp.status_code != 200:
                        failed += 1
                        continue
                    data = iface_resp.json()
                    nic_id = str(data.get("Id", "")).lower()
                    mac = data.get("MACAddress")
                    if not (nic_id and mac):
                        continue

                    if "-" in nic_id:
                        parts = nic_id.split("-")
                        if len(parts) >= 3 and parts[2] != "1":
                            continue

                    if "embedded" in nic_id or "lom" in nic_id:
                        emb_list.append((nic_id, mac.upper()))
                    elif "integrated" in nic_id:
                        int_list.append((nic_id, mac.upper()))
                except Exception:
                    failed += 1

            def extract_nums(s):
                return [int(n) for n in re.findall(r"\d+", s)]

            emb_list.sort(key=lambda x: extract_nums(x[0]))
            int_list.sort(key=lambda x: extract_nums(x[0]))

            for i, (_, mac) in enumerate(emb_list):
                mac_data[f"emb_{i + 1}"] = mac
            for i, (_, mac) in enumerate(int_list):
                mac_data[f"int_{i + 1}"] = mac

            log(f"[SUCCESS] Fetched {len(emb_list) + len(int_list)} NIC MAC address(es) from iDRAC.")
            if failed:
                log(f"[WARNING] {failed} NIC endpoint(s) could not be queried and were skipped.")
        except Exception as e:
            log(f"[ERROR] iDRAC network API failed: {e}")

        for parsed_key, keywords in IDRAC_NIC_COLUMN_KEYWORDS.items():
            col = find_column_index(ws, header_row, keywords)
            if not col:
                continue
            mac = mac_data.get(parsed_key, "N/A")
            write_safe(ws, target_row, col, mac)
            log(f"[UPDATE] {'/'.join(keywords)} -> {mac}")

    # ---------------------------------------------------------
    # PART 2: PHYSICAL DISKS (smart matching by slot + size)
    # ---------------------------------------------------------
    log("\n[API] Fetching physical disk data from iDRAC...")
    drives_list = []

    try:
        resp = requests.get(f"{url}/redfish/v1/Chassis/System.Embedded.1/Drives", auth=auth, verify=False, timeout=10)
        if resp.status_code == 200:
            drives_list = [m.get("@odata.id") for m in resp.json().get("Members", [])]
    except Exception:
        pass

    if not drives_list:
        try:
            resp = requests.get(f"{url}/redfish/v1/Systems/System.Embedded.1/Storage", auth=auth, verify=False, timeout=10)
            if resp.status_code == 200:
                for storage in resp.json().get("Members", []):
                    s_url = storage.get("@odata.id")
                    s_resp = requests.get(f"{url}{s_url}", auth=auth, verify=False, timeout=5)
                    if s_resp.status_code == 200:
                        for drive in s_resp.json().get("Drives", []):
                            drives_list.append(drive.get("@odata.id"))
        except Exception:
            pass

    idrac_disks = []
    for d_url in set(filter(None, drives_list)):
        try:
            d_resp = requests.get(f"{url}{d_url}", auth=auth, verify=False, timeout=5)
            if d_resp.status_code != 200:
                continue
            d_json = d_resp.json()
            serial = str(d_json.get("SerialNumber", "")).strip()
            d_id = d_json.get("Id", "")
            capacity_bytes = d_json.get("CapacityBytes", 0)
            cap_gb = round(capacity_bytes / (1024 ** 3), 2) if capacity_bytes else 0.0

            bay_match = re.search(r"Bay\.(\d+)", d_id, re.IGNORECASE) or re.search(r"Direct\.(\d+)", d_id, re.IGNORECASE)
            bay = int(bay_match.group(1)) if bay_match else 999

            if serial:
                idrac_disks.append({"id": d_id, "serial": serial, "size_gb": cap_gb, "bay": bay, "assigned": False})
        except Exception:
            continue

    if idrac_disks:
        log(f"[SUCCESS] Found {len(idrac_disks)} physical disk(s) via iDRAC.")
    else:
        log("[WARNING] No physical disks found via iDRAC API.")

    disk_start_row = -1
    expected_headers = [f"lf-{server_name.lower()}", server_name.lower()]
    for row in range(1, ws.max_row + 1):
        cell_val = normalize_text(ws.cell(row=row, column=NAME_COL).value)
        if cell_val in expected_headers:
            next_val = normalize_text(ws.cell(row=row + 1, column=NAME_COL).value)
            if "slots" in next_val or "size" in next_val:
                disk_start_row = row + 2
                break

    if disk_start_row == -1:
        log(f"[WARNING] Could not find the Disk section for '{server_name}' in the Excel sheet.")
        return

    log(f"[INFO] Smart-matching disks for '{server_name}'...")
    for row in range(disk_start_row, ws.max_row + 1):
        slot_cell = ws.cell(row=row, column=NAME_COL)
        if type(slot_cell).__name__ == "MergedCell":
            continue

        slot_val = str(slot_cell.value).strip()
        if not slot_val.isdigit():
            break  # end of the numbered slot list

        excel_slot = int(slot_val)
        excel_size_str = str(ws.cell(row=row, column=VAL_COL_FW).value).strip()
        size_match = re.search(r"[\d\.]+", excel_size_str)
        excel_size = float(size_match.group()) if size_match else 0.0

        best_match = None
        min_diff = float("inf")

        # Pass 1: same slot, closest size within tolerance
        for d in idrac_disks:
            if not d["assigned"] and d["bay"] == excel_slot:
                diff = abs(d["size_gb"] - excel_size)
                if diff < min_diff and diff < DISK_SIZE_TOLERANCE_GB:
                    min_diff = diff
                    best_match = d

        # Pass 2: closest size globally
        if not best_match and excel_size > 0:
            min_diff = float("inf")
            for d in idrac_disks:
                if not d["assigned"]:
                    diff = abs(d["size_gb"] - excel_size)
                    if diff < min_diff and diff < DISK_SIZE_TOLERANCE_GB:
                        min_diff = diff
                        best_match = d

        # Pass 3: same slot, any size (fallback)
        if not best_match:
            for d in idrac_disks:
                if not d["assigned"] and d["bay"] == excel_slot:
                    best_match = d
                    break

        if best_match:
            best_match["assigned"] = True
            write_safe(ws, row, VAL_COL_DISK_SN, best_match["serial"])
            log(
                f"[UPDATE] Slot {excel_slot} ({excel_size}G) -> {best_match['id']} "
                f"({best_match['size_gb']}G) -> Serial: {best_match['serial']}"
            )
        else:
            write_safe(ws, row, VAL_COL_DISK_SN, "N/A")
            log(f"[WARNING] Slot {excel_slot} ({excel_size}G) -> N/A (no matching disk found)")


# ============================================================
# ORCHESTRATION
# ============================================================
def run_automation(cfg, log, confirm):
    log("=" * 60)
    log("Starting LinkFlex Inventory Automation")
    log("=" * 60)

    folder = cfg["folder_path"]
    if not os.path.isdir(folder):
        log(f"[FATAL] Folder does not exist: {folder}")
        return False

    excel_path = find_excel_file(folder, log)
    if not excel_path:
        return False

    wb, ws = load_workbook_safe(excel_path, log, confirm)
    if wb is None:
        return False

    all_saved = True
    any_saved = False

    process_firewall(folder, ws, log)
    step_saved = save_excel_safe(wb, excel_path, log, confirm)
    all_saved &= step_saved
    any_saved |= step_saved

    process_proxmox_api(ws, cfg, log)
    step_saved = save_excel_safe(wb, excel_path, log, confirm)
    all_saved &= step_saved
    any_saved |= step_saved

    if cfg.get("run_idrac"):
        process_idrac_api(ws, cfg, log)
        step_saved = save_excel_safe(wb, excel_path, log, confirm)
        all_saved &= step_saved
        any_saved |= step_saved
    else:
        log("\n[INFO] iDRAC processing skipped (disabled).")

    if any_saved:
        export_completed_copy(excel_path, log)
    else:
        log("[WARNING] Nothing was saved - skipping the completed-copy export.")

    log("\n" + "=" * 60)
    log("ALL TASKS COMPLETED" if all_saved else "COMPLETED WITH WARNINGS - see log above")
    log("=" * 60)
    return all_saved


# ============================================================
# GUI
# ============================================================
class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("LinkFlex Inventory Automation")
        self.root.geometry("880x720")
        self.root.minsize(820, 640)

        self.log_queue = queue.Queue()
        self.worker_thread = None
        self.idrac_widgets = []

        self._build_widgets()
        self._poll_log_queue()

    # -------------------- widget builders --------------------
    def _labeled_entry(self, parent, label, default, row):
        var = tk.StringVar(value=default)
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        entry = ttk.Entry(parent, textvariable=var)
        entry.grid(row=row, column=1, columnspan=2, sticky="ew", padx=6, pady=4)
        return var, entry

    def _labeled_password(self, parent, label, default, row):
        var = tk.StringVar(value=default)
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=4)
        entry = ttk.Entry(parent, textvariable=var, show="*")
        entry.grid(row=row, column=1, sticky="ew", padx=6, pady=4)

        show_var = tk.BooleanVar(value=False)

        def toggle():
            entry.config(show="" if show_var.get() else "*")

        ttk.Checkbutton(parent, text="Show", variable=show_var, command=toggle).grid(row=row, column=2, padx=(0, 6))
        return var, entry

    def _build_widgets(self):
        pad = {"padx": 8, "pady": 4}

        # Folder selection
        folder_frame = ttk.LabelFrame(self.root, text="Working Folder")
        folder_frame.pack(fill="x", **pad)
        self.folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_var).pack(
            side="left", fill="x", expand=True, padx=(8, 4), pady=6
        )
        ttk.Button(folder_frame, text="Browse...", command=self._browse_folder).pack(
            side="left", padx=(0, 8), pady=6
        )

        creds_frame = ttk.Frame(self.root)
        creds_frame.pack(fill="x", **pad)

        # Proxmox
        prox_frame = ttk.LabelFrame(creds_frame, text="Proxmox")
        prox_frame.pack(side="left", fill="both", expand=True, padx=(0, 4))
        prox_frame.columnconfigure(1, weight=1)
        self.prox_url, _ = self._labeled_entry(prox_frame, "URL:", "https://192.168.130.31:8006", row=0)
        self.prox_user, _ = self._labeled_entry(prox_frame, "User:", "root", row=1)
        self.prox_pass, _ = self._labeled_password(prox_frame, "Password:", "CnspveAdm@!", row=2)

        # iDRAC
        idrac_frame = ttk.LabelFrame(creds_frame, text="iDRAC (Redfish)")
        idrac_frame.pack(side="left", fill="both", expand=True, padx=(4, 0))
        idrac_frame.columnconfigure(1, weight=1)

        self.idrac_enabled = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            idrac_frame, text="Enable iDRAC processing", variable=self.idrac_enabled, command=self._toggle_idrac_fields
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=6, pady=(6, 2))

        self.idrac_url, w1 = self._labeled_entry(idrac_frame, "URL:", "https://192.168.130.32/", row=1)
        self.idrac_user, w2 = self._labeled_entry(idrac_frame, "User:", "root", row=2)
        self.idrac_pass, w3 = self._labeled_password(idrac_frame, "Password:", "CnsidracAdm@!", row=3)
        self.idrac_widgets.extend([w1, w2, w3])

        ttk.Label(idrac_frame, text="Server Name:").grid(row=4, column=0, sticky="w", padx=6, pady=4)
        self.idrac_server = tk.StringVar(value="PVE1")
        server_combo = ttk.Combobox(
            idrac_frame, textvariable=self.idrac_server, values=["PVE1", "PVE2", "PVE3"], width=16
        )
        server_combo.grid(row=4, column=1, columnspan=2, sticky="ew", padx=6, pady=4)
        self.idrac_widgets.append(server_combo)

        ttk.Label(
            idrac_frame, text="Each PVE host has its own iDRAC IP -\nrun once per server for PVE2 / PVE3.",
            foreground="#777777", justify="left"
        ).grid(row=5, column=0, columnspan=3, sticky="w", padx=6, pady=(0, 6))

        # Controls
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(fill="x", **pad)
        self.run_btn = ttk.Button(btn_frame, text="Run Automation", command=self._on_run)
        self.run_btn.pack(side="left")
        ttk.Button(btn_frame, text="Clear Log", command=self._clear_log).pack(side="left", padx=8)
        self.status_var = tk.StringVar(value="Idle")
        ttk.Label(btn_frame, textvariable=self.status_var).pack(side="right")

        self.progress = ttk.Progressbar(self.root, mode="indeterminate")
        self.progress.pack(fill="x", padx=8, pady=(0, 4))

        # Log
        log_frame = ttk.LabelFrame(self.root, text="Log")
        log_frame.pack(fill="both", expand=True, **pad)
        self.log_text = scrolledtext.ScrolledText(log_frame, state="disabled", wrap="word", font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, padx=4, pady=4)
        self.log_text.tag_config("FATAL", foreground="#ffffff", background="#c0392b")
        self.log_text.tag_config("ERROR", foreground="#c0392b")
        self.log_text.tag_config("WARNING", foreground="#b9770e")
        self.log_text.tag_config("SUCCESS", foreground="#1e8449")
        self.log_text.tag_config("UPDATE", foreground="#1a5276")
        self.log_text.tag_config("INFO", foreground="#555555")

    def _toggle_idrac_fields(self):
        state = "normal" if self.idrac_enabled.get() else "disabled"
        for w in self.idrac_widgets:
            w.configure(state=state)

    # -------------------- actions --------------------
    def _browse_folder(self):
        initial = self.folder_var.get().strip() or os.getcwd()
        if not os.path.isdir(initial):
            initial = os.getcwd()
        path = filedialog.askdirectory(title="Select Working Folder", initialdir=initial)
        if path:
            self.folder_var.set(path)

    def _gather_config(self):
        folder = self.folder_var.get().strip()
        if not folder:
            messagebox.showwarning("Missing Folder", "Please select the working folder first.")
            return None
        if not os.path.isdir(folder):
            messagebox.showerror("Invalid Folder", f"Folder does not exist:\n{folder}")
            return None

        cfg = {
            "folder_path": folder,
            "prox_url": self.prox_url.get().strip(),
            "prox_user": self.prox_user.get().strip(),
            "prox_pass": self.prox_pass.get(),
            "run_idrac": self.idrac_enabled.get(),
        }
        if not cfg["prox_url"] or not cfg["prox_user"]:
            messagebox.showwarning("Missing Proxmox Info", "Please fill in the Proxmox URL and User.")
            return None

        if cfg["run_idrac"]:
            cfg["idrac_url"] = self.idrac_url.get().strip()
            cfg["idrac_user"] = self.idrac_user.get().strip()
            cfg["idrac_pass"] = self.idrac_pass.get()
            cfg["server_name"] = self.idrac_server.get().strip()
            if not all([cfg["idrac_url"], cfg["idrac_user"], cfg["server_name"]]):
                messagebox.showwarning(
                    "Missing iDRAC Info",
                    "Please fill in the iDRAC URL, User and Server Name, or disable iDRAC processing.",
                )
                return None

        return cfg

    def _on_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return
        cfg = self._gather_config()
        if not cfg:
            return
        self._set_running(True)
        self.worker_thread = threading.Thread(target=self._run_worker, args=(cfg,), daemon=True)
        self.worker_thread.start()

    def _run_worker(self, cfg):
        try:
            ok = run_automation(cfg, log=self.log, confirm=self.ask_retry_cancel)
        except Exception as e:
            self.log(f"[FATAL] Unexpected error: {e}")
            self.log(traceback.format_exc())
            ok = False
        self.root.after(0, lambda: self._on_run_finished(ok))

    def _on_run_finished(self, ok):
        self._set_running(False)
        self.status_var.set("Completed" if ok else "Completed with warnings")
        if ok:
            messagebox.showinfo("Done", "Inventory automation completed successfully.")
        else:
            messagebox.showwarning(
                "Done with warnings", "Automation finished but some steps reported errors.\nCheck the log for details."
            )

    def _set_running(self, running):
        self.run_btn.configure(state="disabled" if running else "normal")
        self.status_var.set("Running..." if running else "Idle")
        if running:
            self.progress.start(12)
        else:
            self.progress.stop()

    # -------------------- thread-safe helpers --------------------
    def ask_retry_cancel(self, title, message):
        """Callable from the worker thread; blocks it until the user answers on the main thread."""
        result = {}
        event = threading.Event()

        def show():
            result["value"] = messagebox.askretrycancel(title, message)
            event.set()

        self.root.after(0, show)
        event.wait()
        return result["value"]

    def log(self, message):
        self.log_queue.put(message)

    def _poll_log_queue(self):
        try:
            while True:
                message = self.log_queue.get_nowait()
                self._append_log(message)
        except queue.Empty:
            pass
        self.root.after(150, self._poll_log_queue)

    def _append_log(self, message):
        tag = "INFO"
        for candidate in ("FATAL", "ERROR", "WARNING", "SUCCESS", "UPDATE"):
            if f"[{candidate}]" in message:
                tag = candidate
                break
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n", tag)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")


def run_headless():
    """Entry point when launched by PS Automation as a subprocess (no display
    available - the GUI above can't run). Config comes from PSAUTO_* env vars
    instead of the Tkinter form; the same processing functions run unchanged.

    Unlike the GUI's "pick a working folder" flow (which auto-discovers the
    single .xlsx/firewall .txt already sitting there), this always starts
    from the read-only template next to this script and writes a fresh,
    uniquely-named copy - so concurrent/repeated runs never collide or
    silently overwrite each other or the template itself."""
    def log(message):
        print(message, flush=True)

    def confirm(title, message):
        # No one can answer a dialog in a headless subprocess - log clearly
        # and skip that save step rather than hang forever.
        log(f"[WARNING] {title}: {message.splitlines()[0]} - skipping (make sure the file isn't open elsewhere and re-run if needed).")
        return False

    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, "LinkFlex Inventory-SP_SRV.xlsx")
    if not os.path.exists(template_path):
        log(f"[FATAL] Template not found next to the script: {template_path}")
        sys.exit(1)

    # PS Automation tells us exactly where this run's output belongs (see
    # PSAUTO_RUN_OUTPUT_DIR in server.py's _launch_run) - written there
    # directly rather than next to the script, since a generic sweep-by-
    # extension would also scoop up and relocate the read-only template
    # sitting in this same folder. Falls back to writing next to the script
    # (like before) if that variable isn't set, e.g. someone exporting the
    # other PSAUTO_* vars by hand outside the app.
    output_dir = os.environ.get("PSAUTO_RUN_OUTPUT_DIR", "").strip() or script_dir
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out_path = os.path.join(output_dir, f"LinkFlex_Inventory_Filled_{timestamp}.xlsx")
    shutil.copy2(template_path, out_path)
    log(f"[INFO] Working copy created: {os.path.basename(out_path)}")

    firewall_src = os.environ.get("PSAUTO_FIREWALL_FILE", "").strip()
    firewall_tmp_name = f"ATP_LinkFlex_Firewall_{timestamp}.txt"
    if firewall_src and os.path.isfile(firewall_src):
        shutil.copy2(firewall_src, os.path.join(script_dir, firewall_tmp_name))
        log(f"[INFO] Firewall report received: {os.path.basename(firewall_src)}")
    else:
        log("[WARNING] No firewall report uploaded - firewall MAC/serial fields will be left as-is.")

    prox_cfg = {
        "prox_url": os.environ.get("PSAUTO_PROXMOX_URL", "").strip(),
        "prox_user": os.environ.get("PSAUTO_PROXMOX_USER", "").strip(),
        "prox_pass": os.environ.get("PSAUTO_PROXMOX_PASS", ""),
    }

    # Each PVE host has its OWN iDRAC IP (there's no single shared address),
    # so the app sends a JSON list of {"server": "PVE1", "url": "..."} for
    # however many were checked - zero, one, or all three - login (user/pass)
    # is the same account for all of them. See PSAUTO_IDRAC_TARGETS in
    # server.py's _launch_run and the idrac_targets field in app.js.
    idrac_user = os.environ.get("PSAUTO_IDRAC_USER", "").strip()
    idrac_pass = os.environ.get("PSAUTO_IDRAC_PASS", "")
    idrac_targets_raw = os.environ.get("PSAUTO_IDRAC_TARGETS", "").strip()
    idrac_targets = []
    if idrac_targets_raw:
        try:
            idrac_targets = json.loads(idrac_targets_raw)
        except Exception as e:
            log(f"[ERROR] Could not parse iDRAC target list: {e}")

    wb, ws = load_workbook_safe(out_path, log, confirm)
    if wb is None:
        sys.exit(1)

    any_saved = False
    process_firewall(script_dir, ws, log)
    any_saved |= save_excel_safe(wb, out_path, log, confirm)
    process_proxmox_api(ws, prox_cfg, log)
    any_saved |= save_excel_safe(wb, out_path, log, confirm)

    if not idrac_targets:
        log("\n[INFO] No iDRAC server selected - skipping iDRAC processing.")
    for target in idrac_targets:
        server_name = str(target.get("server", "")).strip()
        idrac_url = str(target.get("url", "")).strip()
        if not server_name or not idrac_url:
            log(f"[WARNING] Skipping incomplete iDRAC target: {target}")
            continue
        idrac_cfg = {
            "idrac_url": idrac_url, "idrac_user": idrac_user, "idrac_pass": idrac_pass,
            "server_name": server_name,
        }
        process_idrac_api(ws, idrac_cfg, log)
        any_saved |= save_excel_safe(wb, out_path, log, confirm)

    if firewall_src:
        try:
            os.remove(os.path.join(script_dir, firewall_tmp_name))
        except Exception:
            pass  # best-effort cleanup of the temp copy - never fails the run over it

    if any_saved:
        log(f"[SUCCESS] Inventory saved: {os.path.basename(out_path)}")
    else:
        log("[WARNING] Nothing was saved - the output file may be incomplete.")
    sys.exit(0 if any_saved else 1)


def main():
    # PSAUTO_PROXMOX_URL is only ever set when PS Automation launches this
    # script - a real interactive/manual run never has it, so this can't
    # accidentally swallow a normal double-click launch.
    if os.environ.get("PSAUTO_PROXMOX_URL"):
        run_headless()
        return

    if not _TKINTER_AVAILABLE:
        print("[FATAL] Tkinter is not available in this Python installation.")
        print("        On Windows, reinstall Python from python.org with the 'tcl/tk' option enabled.")
        sys.exit(1)

    root = tk.Tk()
    try:
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
    except Exception:
        pass
    InventoryApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
