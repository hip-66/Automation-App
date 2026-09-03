import os
import sys
import subprocess
import time
import re
import glob
import traceback

# --- Step 1: Install Libraries ---
def install_and_import(package, import_name=None):
    if import_name is None:
        import_name = package
    try:
        __import__(import_name)
    except ImportError:
        print(f"[INIT] Installing: {package}...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        except:
            print(f"[FATAL] Failed to install {package}.")
            sys.exit(1)

required_libraries = [
    ('openpyxl', 'openpyxl'),
    ('requests', 'requests'),
    ('urllib3', 'urllib3')
]

print("--- Checking libraries ---")
for lib, imp in required_libraries:
    install_and_import(lib, imp)

import openpyxl
import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

NAME_COL = 2        # Column B (Name / Slot)
VAL_COL_FW = 3      # Column C (Firewall MACs / Disk Size)
VAL_COL_DISK_SN = 4 # Column D (Disk Serial Number)
VAL_COL_VM_1 = 6    # Column F (Data)
VAL_COL_VM_2 = 7    # Column G (MGMT)

def get_validated_input(prompt, allowed_options=None):
    print("\n" + "-" * 60)
    print(prompt)
    if allowed_options:
        print(f"Options: {allowed_options}")
    print("-" * 60)

    while True:
        user_in = input(">> ").strip()
        
        if user_in.lower() == 'exit':
            sys.exit(0)

        if any("\u0590" <= c <= "\u05FF" for c in user_in):
            print("[ERROR] Hebrew detected. Please use English only.")
            continue

        if allowed_options:
            if user_in.lower() not in [x.lower() for x in allowed_options]:
                print(f"[ERROR] Invalid choice. Options: {allowed_options}")
                continue

        if not user_in:
            print("[ERROR] Input cannot be empty.")
            continue

        return user_in.replace('"', '')

def normalize_text(text):
    if not text: return ""
    return str(text).lower().replace(" ", "").strip()

def save_excel_safe(workbook, path):
    try:
        workbook.save(path)
        print(f"[SUCCESS] Excel saved successfully.")
        return True
    except PermissionError:
        print(f"\n[FATAL ERROR] Excel is OPEN! Please close it and try again.")
        return False
    except Exception as e:
        print(f"[ERROR] Save failed: {e}")
        return False

def find_column_index(sheet, row_idx, keywords):
    for col in range(1, sheet.max_column + 1):
        val = normalize_text(sheet.cell(row=row_idx, column=col).value)
        if all(k.lower() in val for k in keywords):
            return col
    return None

def write_safe(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    if type(cell).__name__ == 'MergedCell':
        return False
    cell.value = value
    return True

# --- FIREWALL ---
def process_firewall(folder_path, excel_ws):
    print("\n--- Processing Firewall ---")
    txt_files = glob.glob(os.path.join(folder_path, "ATP_LinkFlex_Firewall*.txt"))
    if not txt_files:
        print("[ERROR] No Firewall text file found.")
        return

    latest_file = max(txt_files, key=os.path.getctime)
    mac_mapping = {}
    serial_number = "N/A"
    
    try:
        with open(latest_file, 'r', encoding='utf-8') as f:
            content = f.read()
            if "----- INVENTORY -----" in content:
                inv_section = content.split("----- INVENTORY -----")[1]
                
                serial_match = re.search(r'Serial Number:\s*([A-Za-z0-9-]+)', inv_section)
                if serial_match:
                    serial_number = serial_match.group(1).strip()
                
                for line in inv_section.splitlines():
                    match = re.search(r'^\s*([a-zA-Z0-9]+):\s*([0-9a-fA-F:]+)', line)
                    if match:
                        mac_mapping[match.group(1)] = match.group(2).lower()
    except Exception as e:
        print(f"[ERROR] Reading Firewall file failed: {e}")
        return

    target_ports = {
        'port1': 'FW MAC PORT 1', 'port2': 'FW MAC PORT 2', 'port3': 'FW MAC PORT 3',
        'port4': 'FW MAC PORT 4', 'port5': 'FW MAC PORT 5', 'port6': 'FW MAC PORT 6',
        'port7': 'FW MAC PORT 7', 'port8': 'FW MAC PORT 8', 'port9': 'FW MAC PORT 9',
        'port10': 'FW MAC PORT 10', 'port11': 'FW MAC PORT 11', 'port12': 'FW MAC PORT 12',
        'port15': 'FW MAC PORT 15', 'port16': 'FW MAC PORT 16',
        'x1': 'FW MAC PORT x1', 'x2': 'FW MAC PORT x2', 'x3': 'FW MAC PORT x3',
        'x6': 'FW MAC PORT x6', 'x8': 'FW MAC PORT x8'
    }

    norm_targets = {k: normalize_text(v) for k, v in target_ports.items()}
    
    for row in range(1, excel_ws.max_row + 1):
        cell_obj = excel_ws.cell(row=row, column=NAME_COL)
        if type(cell_obj).__name__ == 'MergedCell': continue
        
        val = normalize_text(cell_obj.value)
        
        if val == "fw":
            write_safe(excel_ws, row, VAL_COL_FW, serial_number)
            print(f"[UPDATE] Firewall Serial Number -> {serial_number}")
            continue

        for key, target in norm_targets.items():
            if target == val:
                write_safe(excel_ws, row, VAL_COL_FW, mac_mapping.get(key, "N/A"))
                print(f"[UPDATE] {target} -> {mac_mapping.get(key, 'N/A')}")

# --- PROXMOX API ---
def process_proxmox_api(excel_ws, config):
    print("\n--- Processing Proxmox (via API) ---")
    url = config['prox_url']
    if not url.startswith("http"): url = "https://" + url
    
    user = config['prox_user']
    if '@' not in user:
        user += '@pam' 

    print("[API] Authenticating to Proxmox...")
    auth_url = f"{url}/api2/json/access/ticket"
    try:
        resp = requests.post(auth_url, data={'username': user, 'password': config['prox_pass']}, verify=False, timeout=10)
        resp.raise_for_status()
        auth_data = resp.json()['data']
        headers = {
            'CSRFPreventionToken': auth_data['CSRFPreventionToken'],
            'Cookie': f"PVEAuthCookie={auth_data['ticket']}"
        }
        print("[SUCCESS] Connected to Proxmox API.")
    except Exception as e:
        print(f"[ERROR] Proxmox API Login Failed (Check credentials/URL).")
        return

    print("[API] Fetching VM List...")
    try:
        res_resp = requests.get(f"{url}/api2/json/cluster/resources?type=vm", headers=headers, verify=False, timeout=10)
        res_resp.raise_for_status()
        vms = res_resp.json()['data']
        vm_dict = {str(vm.get('name', '')).lower(): vm for vm in vms}
    except Exception as e:
        print(f"[ERROR] Failed to fetch VMs: {e}")
        return

    start_row = -1
    end_row = excel_ws.max_row
    
    for row in range(1, excel_ws.max_row + 1):
        val = str(excel_ws.cell(row=row, column=NAME_COL).value)
        if "PVE VM's" in val:
            start_row = row + 2 
            break
            
    if start_row == -1: return print("[ERROR] 'PVE VM's' section not found.")

    for row in range(start_row, excel_ws.max_row + 1):
        val = str(excel_ws.cell(row=row, column=NAME_COL).value)
        if "Hypervisor Servers" in val:
            end_row = row - 1
            break

    for row in range(start_row, end_row + 1):
        name_cell = excel_ws.cell(row=row, column=NAME_COL)
        if type(name_cell).__name__ == 'MergedCell':
            continue

        vm_name_raw = name_cell.value
        if not vm_name_raw or str(vm_name_raw).lower() == 'none': continue

        vm_name = str(vm_name_raw).lower()
        
        dest_cell = excel_ws.cell(row=row, column=VAL_COL_VM_1)
        if type(dest_cell).__name__ == 'MergedCell':
            continue
            
        if vm_name in vm_dict:
            node = vm_dict[vm_name]['node']
            vmid = vm_dict[vm_name]['vmid']
            print(f"[CHECK] Querying VM: {vm_name_raw} (ID: {vmid})")
            
            config_url = f"{url}/api2/json/nodes/{node}/qemu/{vmid}/config"
            macs = []
            try:
                conf_resp = requests.get(config_url, headers=headers, verify=False, timeout=5)
                if conf_resp.status_code == 200:
                    vm_config = conf_resp.json().get('data', {})
                    for i in range(4):
                        net_str = vm_config.get(f'net{i}', '')
                        match = re.search(r'([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})', net_str)
                        if match:
                            macs.append(match.group(0).lower())
            except Exception as e:
                print(f"[DEBUG] API error for VM {vm_name_raw}: {e}")

            if len(macs) == 0:
                write_safe(excel_ws, row, VAL_COL_VM_1, "N/A")
                print(f"[UPDATE] {vm_name_raw} -> N/A")
            elif len(macs) == 1:
                write_safe(excel_ws, row, VAL_COL_VM_1, macs[0])
                print(f"[UPDATE] {vm_name_raw} -> {macs[0]}")
            else:
                mac_data = f"Data - {macs[0]}"
                mac_mgmt = f"MGMT - {macs[1]}"
                
                write_safe(excel_ws, row, VAL_COL_VM_1, mac_data)
                write_safe(excel_ws, row, VAL_COL_VM_2, mac_mgmt)
                print(f"[UPDATE] {vm_name_raw} -> {mac_data} | {mac_mgmt}")
        else:
            if len(vm_name) > 2 and "slots" not in vm_name:
                write_safe(excel_ws, row, VAL_COL_VM_1, "N/A")

# --- iDRAC API (Redfish) ---
def process_idrac_api(excel_ws, config):
    print("\n--- Processing iDRAC (via Redfish API) ---")
    url = config['idrac_url']
    if not url.startswith("http"): url = "https://" + url

    auth = (config['idrac_user'], config['idrac_pass'])
    
    # ---------------------------------------------------------
    # PART 1: NETWORK INTERFACES (NICs)
    # ---------------------------------------------------------
    target_row = -1
    header_row = -1
    for row in range(1, 200):
        if "Hypervisor Servers" in str(excel_ws.cell(row=row, column=NAME_COL).value):
            header_row = row + 1
            break
            
    if header_row != -1:
        for row in range(header_row + 1, excel_ws.max_row):
            if normalize_text(excel_ws.cell(row=row, column=NAME_COL).value) == normalize_text(config['server_name']):
                target_row = row
                break
                
    if target_row == -1:
        print(f"[ERROR] Server '{config['server_name']}' not found in Hypervisor section.")
    else:
        print("[API] Fetching Network Data from iDRAC...")
        base_redfish_url = f"{url}/redfish/v1/Systems/System.Embedded.1/EthernetInterfaces"
        
        emb_list = []
        int_list = []
        
        try:
            resp = requests.get(base_redfish_url, auth=auth, verify=False, timeout=10)
            resp.raise_for_status()
            interfaces = resp.json().get('Members', [])
            
            for iface in interfaces:
                iface_url = iface.get('@odata.id')
                if iface_url:
                    try:
                        iface_resp = requests.get(f"{url}{iface_url}", auth=auth, verify=False, timeout=5)
                        if iface_resp.status_code == 200:
                            data = iface_resp.json()
                            nic_id = str(data.get('Id', '')).lower()
                            mac = data.get('MACAddress')
                            
                            if nic_id and mac:
                                if '-' in nic_id:
                                    parts = nic_id.split('-')
                                    if len(parts) >= 3 and parts[2] != '1':
                                        continue
                                
                                if "embedded" in nic_id or "lom" in nic_id:
                                    emb_list.append((nic_id, mac.upper()))
                                elif "integrated" in nic_id:
                                    int_list.append((nic_id, mac.upper()))
                    except: pass
                    
            def extract_nums(s):
                return [int(n) for n in re.findall(r'\d+', s)]
                
            emb_list.sort(key=lambda x: extract_nums(x[0]))
            int_list.sort(key=lambda x: extract_nums(x[0]))
            
            mac_data = {}
            for i, (nic_id, mac) in enumerate(emb_list):
                mac_data[f"emb_{i+1}"] = mac
            for i, (nic_id, mac) in enumerate(int_list):
                mac_data[f"int_{i+1}"] = mac

            print("[SUCCESS] Fetched MAC addresses from iDRAC.")
        except Exception as e:
            print(f"[ERROR] iDRAC Network API failed: {e}")

        api_mapping = {
            "emb_1": ["NIC1", "Emb"], "emb_2": ["NIC2", "Emb"],
            "int_1": ["NIC1", "Int"], "int_2": ["NIC2", "Int"],
            "int_3": ["NIC3", "Int"], "int_4": ["NIC4", "Int"]
        }

        for parsed_key, keywords in api_mapping.items():
            col = find_column_index(excel_ws, header_row, keywords)
            if not col: continue
            
            mac = mac_data.get(parsed_key, "N/A")
            write_safe(excel_ws, target_row, col, mac)
            print(f"[UPDATE] iDRAC NIC -> {mac}")

    # ---------------------------------------------------------
    # PART 2: PHYSICAL DISKS (Smart Matching by Slot + EXACT Size)
    # ---------------------------------------------------------
    print("\n[API] Fetching Physical Disks Data from iDRAC...")
    drives_list = []
    
    try:
        resp = requests.get(f"{url}/redfish/v1/Chassis/System.Embedded.1/Drives", auth=auth, verify=False, timeout=10)
        if resp.status_code == 200:
            for member in resp.json().get('Members', []):
                drives_list.append(member.get('@odata.id'))
    except: pass
    
    if not drives_list:
        try:
            resp = requests.get(f"{url}/redfish/v1/Systems/System.Embedded.1/Storage", auth=auth, verify=False, timeout=10)
            if resp.status_code == 200:
                for storage in resp.json().get('Members', []):
                    s_url = storage.get('@odata.id')
                    s_resp = requests.get(f"{url}{s_url}", auth=auth, verify=False, timeout=5)
                    if s_resp.status_code == 200:
                        for drive in s_resp.json().get('Drives', []):
                            drives_list.append(drive.get('@odata.id'))
        except: pass

    # Extract detailed disk information
    idrac_disks = []
    for d_url in set(drives_list):
        try:
            d_resp = requests.get(f"{url}{d_url}", auth=auth, verify=False, timeout=5)
            if d_resp.status_code == 200:
                d_json = d_resp.json()
                serial = str(d_json.get('SerialNumber', '')).strip()
                d_id = d_json.get('Id', '')
                capacity_bytes = d_json.get('CapacityBytes', 0)
                
                cap_gb = round(capacity_bytes / (1024**3), 2) if capacity_bytes else 0.0

                bay_match = re.search(r'Bay\.(\d+)', d_id, re.IGNORECASE)
                if not bay_match:
                    bay_match = re.search(r'Direct\.(\d+)', d_id, re.IGNORECASE)
                    
                bay = int(bay_match.group(1)) if bay_match else 999
                
                if serial:
                    idrac_disks.append({
                        'id': d_id,
                        'serial': serial,
                        'size_gb': cap_gb,
                        'bay': bay,
                        'assigned': False # Lock mechanism
                    })
        except: pass

    if idrac_disks:
        print(f"[SUCCESS] Found {len(idrac_disks)} Physical Disks in iDRAC.")
    else:
        print("[WARNING] No Physical Disks found via API.")

    # Find the Disk Section in Excel
    disk_start_row = -1
    expected_headers = [f"lf-{config['server_name'].lower()}", config['server_name'].lower()]
    
    for row in range(1, excel_ws.max_row + 1):
        cell_val = normalize_text(excel_ws.cell(row=row, column=NAME_COL).value)
        if cell_val in expected_headers:
            next_val = normalize_text(excel_ws.cell(row=row+1, column=NAME_COL).value)
            if "slots" in next_val or "size" in next_val:
                disk_start_row = row + 2
                break

    if disk_start_row != -1:
        print(f"[INFO] Smart-Matching Disks in Excel for '{config['server_name']}'...")
        for row in range(disk_start_row, excel_ws.max_row + 1):
            slot_cell = excel_ws.cell(row=row, column=NAME_COL)
            if type(slot_cell).__name__ == 'MergedCell': continue
            
            slot_val = str(slot_cell.value).strip()
            if not slot_val.isdigit():
                break # We reached the end of the numbered slots
                
            excel_slot = int(slot_val)
            excel_size_str = str(excel_ws.cell(row=row, column=VAL_COL_FW).value).strip()
            
            size_match = re.search(r'[\d\.]+', excel_size_str)
            excel_size = float(size_match.group()) if size_match else 0.0
            
            best_match = None
            min_diff = float('inf')
            
            # Pass 1: Match by Slot AND closest Size (Tolerance up to 10GB for logic differences)
            for d in idrac_disks:
                if not d['assigned'] and d['bay'] == excel_slot:
                    diff = abs(d['size_gb'] - excel_size)
                    if diff < min_diff and diff < 10.0:
                        min_diff = diff
                        best_match = d
            
            # Pass 2: Match by Closest Size globally (If slot didn't work)
            if not best_match and excel_size > 0:
                min_diff = float('inf')
                for d in idrac_disks:
                    if not d['assigned']:
                        diff = abs(d['size_gb'] - excel_size)
                        if diff < min_diff and diff < 10.0:
                            min_diff = diff
                            best_match = d
                            
            # Pass 3: Match strictly by Slot (Fallback)
            if not best_match:
                for d in idrac_disks:
                    if not d['assigned'] and d['bay'] == excel_slot:
                        best_match = d
                        break
            
            # Write and Lock
            if best_match:
                best_match['assigned'] = True
                write_safe(excel_ws, row, VAL_COL_DISK_SN, best_match['serial'])
                print(f"[UPDATE] Excel Slot {excel_slot} (Size: {excel_size}G) -> Matched iDRAC {best_match['id']} (Size: {best_match['size_gb']}G) -> Serial: {best_match['serial']}")
            else:
                write_safe(excel_ws, row, VAL_COL_DISK_SN, "N/A")
                print(f"[WARNING] Excel Slot {excel_slot} (Size: {excel_size}G) -> N/A (No matching disk found)")
    else:
        print(f"[WARNING] Could not find the Disk section for '{config['server_name']}' in Excel.")

# --- MAIN ---
def main():
    print("======================================================")
    print("--- Inventory Automation v20.0 (Precision Disks) ---")
    print("======================================================")
    print("Tip: Type 'exit' to quit anytime.")
    
    cfg = {}
    cfg['folder_path'] = get_validated_input("Paste FOLDER PATH:")
    if not os.path.exists(cfg['folder_path']):
        print("[FATAL] Path does not exist."); return

    cfg['prox_url'] = get_validated_input("Proxmox URL (e.g. 192.168.1.10:8006):")
    cfg['prox_user'] = get_validated_input("Proxmox User (e.g. root or root@pam):")
    cfg['prox_pass'] = get_validated_input("Proxmox Password:")
    
    do_idrac = get_validated_input("Run iDRAC? (y/n):", allowed_options=['y', 'n'])
    cfg['run_idrac'] = (do_idrac.lower() == 'y')
    
    if cfg['run_idrac']:
        cfg['idrac_url'] = get_validated_input("iDRAC URL (e.g. 192.168.1.20):")
        cfg['idrac_user'] = get_validated_input("iDRAC User:")
        cfg['idrac_pass'] = get_validated_input("iDRAC Password:")
        cfg['server_name'] = get_validated_input("Server Name in Excel (e.g. PVE1):")

    print("\n[INFO] Starting API Automation... This will be FAST!")
    
    xls_path = os.path.join(cfg['folder_path'], "LinkFlex Inventory.xlsx")
    if not os.path.exists(xls_path):
        found = glob.glob(os.path.join(cfg['folder_path'], "*.xlsx"))
        if found: xls_path = found[0]
        else: return print("No Excel found.")

    try:
        wb = openpyxl.load_workbook(xls_path)
        ws = wb.active
    except: return print("Excel is OPEN. Close it!")

    process_firewall(cfg['folder_path'], ws)
    save_excel_safe(wb, xls_path)

    process_proxmox_api(ws, cfg)
    save_excel_safe(wb, xls_path)

    if cfg['run_idrac']:
        process_idrac_api(ws, cfg)
        save_excel_safe(wb, xls_path)

    print("\n" + "="*50)
    print("ALL TASKS COMPLETED SUCCESSFULLY.")
    print("="*50)
    input("Press Enter to close...")

if __name__ == "__main__":
    try: main()
    except Exception as e:
        print(f"Crash: {e}")
        traceback.print_exc()
        input()